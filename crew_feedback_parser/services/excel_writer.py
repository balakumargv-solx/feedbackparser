"""
Excel writer service for crew feedback parsing system.
Manages Excel file creation, data appending, and formatting.
"""
import os
from datetime import datetime
from pathlib import Path
from typing import Optional, List

import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, NamedStyle
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.worksheet.datavalidation import DataValidation

from ..models.feedback_data import FeedbackData, ProcessingResult


class ExcelWriter:
    """
    Manages Excel file operations for crew feedback data storage and processing tracking.
    
    Creates and maintains a single Excel file with two sheets:
    - "Feedback_Data": Contains extracted feedback data
    - "Processing_Log": Tracks file processing status
    """
    
    # Column headers for Feedback_Data sheet
    FEEDBACK_COLUMNS = [
        "vessel",
        "crew_name", 
        "crew_rank",
        "safer_with_sos",
        "fatigue_monitoring_prevention",
        "geofence_awareness",
        "heat_stress_alerts_change",
        "work_rest_hour_monitoring",
        "ptw_system_improvement",
        "paperwork_error_reduction",
        "noise_exposure_monitoring",
        "activity_tracking_awareness",
        "fall_detection_confidence",
        "feature_preference"
    ]
    
    # Column headers for Processing_Log sheet
    LOG_COLUMNS = [
        "file_name",
        "status",
        "error_message",
        "processing_timestamp"
    ]
    
    def __init__(self, file_path: str):
        """
        Initialize Excel writer with target file path.
        
        Args:
            file_path (str): Path to Excel file to create or load
        """
        self.file_path = Path(file_path)
        self.workbook: Optional[Workbook] = None
        self.feedback_sheet: Optional[Worksheet] = None
        self.log_sheet: Optional[Worksheet] = None
        
    def create_or_load_workbook(self) -> None:
        """
        Create new Excel workbook or load existing one with proper sheet structure.
        Sets up both "Feedback_Data" and "Processing_Log" sheets with headers.
        """
        if self.file_path.exists():
            # Load existing workbook
            self.workbook = openpyxl.load_workbook(self.file_path)
            
            # Get or create Feedback_Data sheet
            if "Feedback_Data" in self.workbook.sheetnames:
                self.feedback_sheet = self.workbook["Feedback_Data"]
            else:
                self.feedback_sheet = self.workbook.create_sheet("Feedback_Data")
                self._setup_feedback_sheet_headers()
            
            # Get or create Processing_Log sheet
            if "Processing_Log" in self.workbook.sheetnames:
                self.log_sheet = self.workbook["Processing_Log"]
            else:
                self.log_sheet = self.workbook.create_sheet("Processing_Log")
                self._setup_log_sheet_headers()
                
            # Remove default sheet if it exists and is empty
            if "Sheet" in self.workbook.sheetnames and len(self.workbook.sheetnames) > 1:
                default_sheet = self.workbook["Sheet"]
                if default_sheet.max_row == 1 and default_sheet.max_column == 1:
                    self.workbook.remove(default_sheet)
        else:
            # Create new workbook
            self.workbook = Workbook()
            
            # Remove default sheet
            if "Sheet" in self.workbook.sheetnames:
                self.workbook.remove(self.workbook["Sheet"])
            
            # Create Feedback_Data sheet
            self.feedback_sheet = self.workbook.create_sheet("Feedback_Data")
            self._setup_feedback_sheet_headers()
            
            # Create Processing_Log sheet
            self.log_sheet = self.workbook.create_sheet("Processing_Log")
            self._setup_log_sheet_headers()
            
            # Ensure parent directory exists
            self.file_path.parent.mkdir(parents=True, exist_ok=True)
    
    def _setup_feedback_sheet_headers(self) -> None:
        """
        Set up column headers for the Feedback_Data sheet.
        """
        if not self.feedback_sheet:
            raise ValueError("Feedback sheet not initialized")
            
        # Add headers if sheet is empty
        if self.feedback_sheet.max_row == 1 and self.feedback_sheet.max_column == 1:
            for col_idx, header in enumerate(self.FEEDBACK_COLUMNS, 1):
                cell = self.feedback_sheet.cell(row=1, column=col_idx, value=header)
                # Apply header formatting
                cell.font = Font(bold=True)
                cell.alignment = Alignment(horizontal='center')
                cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
            
            # Apply formatting and data validation after headers are set
            self._apply_feedback_sheet_formatting()
    
    def _setup_log_sheet_headers(self) -> None:
        """
        Set up column headers for the Processing_Log sheet.
        """
        if not self.log_sheet:
            raise ValueError("Log sheet not initialized")
            
        # Add headers if sheet is empty
        if self.log_sheet.max_row == 1 and self.log_sheet.max_column == 1:
            for col_idx, header in enumerate(self.LOG_COLUMNS, 1):
                cell = self.log_sheet.cell(row=1, column=col_idx, value=header)
                # Apply header formatting
                cell.font = Font(bold=True)
                cell.alignment = Alignment(horizontal='center')
                cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
            
            # Apply formatting after headers are set
            self._apply_log_sheet_formatting()
    
    def _apply_feedback_sheet_formatting(self) -> None:
        """
        Apply proper data types and formatting to Feedback_Data sheet columns.
        """
        if not self.feedback_sheet:
            raise ValueError("Feedback sheet not initialized")
        
        # Set column widths for better readability
        column_widths = {
            'A': 15,  # vessel
            'B': 20,  # crew_name
            'C': 15,  # crew_rank
            'D': 18,  # safer_with_sos
            'E': 25,  # fatigue_monitoring_prevention
            'F': 18,  # geofence_awareness
            'G': 25,  # heat_stress_alerts_change
            'H': 25,  # work_rest_hour_monitoring
            'I': 22,  # ptw_system_improvement
            'J': 25,  # paperwork_error_reduction
            'K': 25,  # noise_exposure_monitoring
            'L': 25,  # activity_tracking_awareness
            'M': 25,  # fall_detection_confidence
            'N': 30,  # feature_preference
        }
        
        for col_letter, width in column_widths.items():
            self.feedback_sheet.column_dimensions[col_letter].width = width
        
        # Add data validation for rating columns (1-5 integers)
        # Rating columns are D through M (columns 4-13)
        rating_validation = DataValidation(
            type="whole",
            operator="between",
            formula1=1,
            formula2=5,
            showErrorMessage=True,
            errorTitle="Invalid Rating",
            error="Rating must be an integer between 1 and 5"
        )
        
        # Apply validation to rating columns for a reasonable range of rows (1000 rows)
        for col_idx in range(4, 14):  # Columns D through M
            col_letter = get_column_letter(col_idx)
            rating_validation.add(f"{col_letter}2:{col_letter}1000")
        
        self.feedback_sheet.add_data_validation(rating_validation)
        
        # Freeze the header row
        self.feedback_sheet.freeze_panes = "A2"
    
    def _apply_log_sheet_formatting(self) -> None:
        """
        Apply proper data types and formatting to Processing_Log sheet columns.
        """
        if not self.log_sheet:
            raise ValueError("Log sheet not initialized")
        
        # Set column widths
        column_widths = {
            'A': 30,  # file_name
            'B': 12,  # status
            'C': 50,  # error_message
            'D': 20,  # processing_timestamp
        }
        
        for col_letter, width in column_widths.items():
            self.log_sheet.column_dimensions[col_letter].width = width
        
        # Add data validation for status column
        status_validation = DataValidation(
            type="list",
            formula1='"pass,fail,error"',
            showErrorMessage=True,
            errorTitle="Invalid Status",
            error="Status must be one of: pass, fail, error"
        )
        
        # Apply validation to status column for a reasonable range of rows
        status_validation.add("B2:B1000")
        self.log_sheet.add_data_validation(status_validation)
        
        # Freeze the header row
        self.log_sheet.freeze_panes = "A2"
    
    def apply_formatting_to_existing_data(self) -> None:
        """
        Apply formatting to existing data in both sheets.
        Useful when loading an existing workbook that may not have proper formatting.
        """
        if self.feedback_sheet:
            self._apply_feedback_sheet_formatting()
        
        if self.log_sheet:
            self._apply_log_sheet_formatting()
    
    def format_rating_cells(self, start_row: int, end_row: int) -> None:
        """
        Apply specific formatting to rating cells in a range of rows.
        
        Args:
            start_row (int): Starting row number
            end_row (int): Ending row number
        """
        if not self.feedback_sheet:
            raise ValueError("Feedback sheet not initialized")
        
        # Rating columns are D through M (columns 4-13)
        for row in range(start_row, end_row + 1):
            for col_idx in range(4, 14):
                cell = self.feedback_sheet.cell(row=row, column=col_idx)
                # Center align rating values
                cell.alignment = Alignment(horizontal='center')
                
                # Validate that the value is an integer between 1-5
                if cell.value is not None:
                    try:
                        rating = int(cell.value)
                        if not (1 <= rating <= 5):
                            # Highlight invalid ratings in red
                            cell.fill = PatternFill(start_color="FFCCCC", end_color="FFCCCC", fill_type="solid")
                    except (ValueError, TypeError):
                        # Highlight non-integer values in red
                        cell.fill = PatternFill(start_color="FFCCCC", end_color="FFCCCC", fill_type="solid")
    
    def save_workbook(self) -> None:
        """
        Save the workbook to file.
        """
        if not self.workbook:
            raise ValueError("Workbook not initialized")
            
        self.workbook.save(self.file_path)
    
    def append_feedback_data(self, feedback_data: FeedbackData) -> None:
        """
        Add new row to Feedback_Data sheet with extracted data.
        
        Args:
            feedback_data (FeedbackData): The feedback data to append
        """
        if not self.feedback_sheet:
            raise ValueError("Feedback sheet not initialized")
        
        # Find next empty row
        next_row = self.feedback_sheet.max_row + 1
        
        # Prepare data row in correct column order
        data_row = [
            feedback_data.vessel,
            feedback_data.crew_name,
            feedback_data.crew_rank,
            feedback_data.safer_with_sos,
            feedback_data.fatigue_monitoring_prevention,
            feedback_data.geofence_awareness,
            feedback_data.heat_stress_alerts_change,
            feedback_data.work_rest_hour_monitoring,
            feedback_data.ptw_system_improvement,
            feedback_data.paperwork_error_reduction,
            feedback_data.noise_exposure_monitoring,
            feedback_data.activity_tracking_awareness,
            feedback_data.fall_detection_confidence,
            feedback_data.feature_preference
        ]
        
        # Add data to row
        for col_idx, value in enumerate(data_row, 1):
            self.feedback_sheet.cell(row=next_row, column=col_idx, value=value)
        
        # Apply formatting to the newly added row
        self.format_rating_cells(next_row, next_row)
    
    def update_processing_log(self, processing_result: ProcessingResult) -> None:
        """
        Update Processing_Log sheet with file processing status.
        
        Args:
            processing_result (ProcessingResult): The processing result to log
        """
        if not self.log_sheet:
            raise ValueError("Log sheet not initialized")
        
        # Find next empty row
        next_row = self.log_sheet.max_row + 1
        
        # Prepare log row data
        log_row = [
            processing_result.file_name,
            processing_result.status,
            processing_result.error_message or "",
            processing_result.processing_timestamp
        ]
        
        # Add data to row
        for col_idx, value in enumerate(log_row, 1):
            self.log_sheet.cell(row=next_row, column=col_idx, value=value)
    
    def append_feedback_and_log(self, processing_result: ProcessingResult) -> None:
        """
        Convenience method to append both feedback data and processing log entry.
        
        Args:
            processing_result (ProcessingResult): The processing result containing data and status
        """
        # Always log the processing result
        self.update_processing_log(processing_result)
        
        # Only append feedback data if processing was successful and data exists
        if processing_result.is_successful() and processing_result.has_data():
            self.append_feedback_data(processing_result.data)
    
    def get_feedback_data_count(self) -> int:
        """
        Get the number of feedback data rows (excluding header).
        
        Returns:
            int: Number of data rows in Feedback_Data sheet
        """
        if not self.feedback_sheet:
            return 0
        return max(0, self.feedback_sheet.max_row - 1)
    
    def get_processing_log_count(self) -> int:
        """
        Get the number of processing log entries (excluding header).
        
        Returns:
            int: Number of log entries in Processing_Log sheet
        """
        if not self.log_sheet:
            return 0
        return max(0, self.log_sheet.max_row - 1)
    
    def get_processing_summary(self) -> dict:
        """
        Get summary statistics from the processing log.
        
        Returns:
            dict: Summary with counts of pass, fail, error statuses
        """
        if not self.log_sheet:
            return {"pass": 0, "fail": 0, "error": 0, "total": 0}
        
        summary = {"pass": 0, "fail": 0, "error": 0, "total": 0}
        
        # Skip header row, start from row 2
        for row in range(2, self.log_sheet.max_row + 1):
            status_cell = self.log_sheet.cell(row=row, column=2)  # Status is column 2
            if status_cell.value:
                status = str(status_cell.value).lower()
                if status in summary:
                    summary[status] += 1
                summary["total"] += 1
        
        return summary

    def close_workbook(self) -> None:
        """
        Close the workbook and clean up resources.
        """
        if self.workbook:
            self.workbook.close()
            self.workbook = None
            self.feedback_sheet = None
            self.log_sheet = None